"""
Report export utilities.
"""

import io
from datetime import date
from decimal import Decimal

from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


class PDFExporter:
    """Export reports to PDF using reportlab."""
    
    @staticmethod
    def export_daily_sales(data: list, start_date: date, end_date: date) -> bytes:
        """Export daily sales report to PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            alignment=1,
            spaceAfter=30,
        )
        elements.append(Paragraph(f"Daily Sales Report", title_style))
        elements.append(Paragraph(f"{start_date} to {end_date}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Table
        table_data = [['Date', 'Total Sales', 'Orders', 'Avg Order', 'Tax', 'Discount']]
        for row in data:
            table_data.append([
                str(row['date']),
                f"₹{row['total_sales']:,.2f}",
                str(row['total_orders']),
                f"₹{row['average_order_value']:,.2f}",
                f"₹{row['total_tax']:,.2f}",
                f"₹{row['total_discount']:,.2f}",
            ])
        
        # Totals
        totals = {
            'sales': sum(r['total_sales'] for r in data),
            'orders': sum(r['total_orders'] for r in data),
            'tax': sum(r['total_tax'] for r in data),
            'discount': sum(r['total_discount'] for r in data),
        }
        avg_total = totals['sales'] / totals['orders'] if totals['orders'] else 0
        table_data.append([
            'TOTAL',
            f"₹{totals['sales']:,.2f}",
            str(totals['orders']),
            f"₹{avg_total:,.2f}",
            f"₹{totals['tax']:,.2f}",
            f"₹{totals['discount']:,.2f}",
        ])
        
        table = Table(table_data, colWidths=[1.2*inch, 1.2*inch, 0.8*inch, 1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        
        doc.build(elements)
        return buffer.getvalue()
    
    @staticmethod
    def export_product_sales(data: list, start_date: date, end_date: date) -> bytes:
        """Export product sales report to PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        elements.append(Paragraph("Product Sales Report", styles['Heading1']))
        elements.append(Paragraph(f"{start_date} to {end_date}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Table
        table_data = [['Product', 'Category', 'Qty Sold', 'Sales', '%']]
        for row in data:
            table_data.append([
                row['product_name'][:30],
                row['category_name'][:20],
                f"{row['quantity_sold']:.0f}",
                f"₹{row['total_sales']:,.2f}",
                f"{row['percentage']:.1f}%",
            ])
        
        table = Table(table_data, colWidths=[2*inch, 1.5*inch, 0.8*inch, 1.2*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        
        doc.build(elements)
        return buffer.getvalue()
    
    @staticmethod
    def export_session_summary(data: list, start_date: date, end_date: date) -> bytes:
        """Export session summary to PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []
        
        elements.append(Paragraph("Session Summary Report", styles['Heading1']))
        elements.append(Paragraph(f"{start_date} to {end_date}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Table
        table_data = [['Terminal', 'User', 'Opened', 'Closed', 'Opening', 'Sales', 'Orders', 'Closing', 'Diff']]
        for row in data:
            closed = row['closed_at'].strftime('%d/%m %H:%M') if row['closed_at'] else '-'
            diff = f"₹{row['difference']:,.2f}" if row['difference'] is not None else '-'
            closing = f"₹{row['closing_balance']:,.2f}" if row['closing_balance'] else '-'
            table_data.append([
                row['terminal_name'][:15],
                row['user_name'][:12],
                row['opened_at'].strftime('%d/%m %H:%M'),
                closed,
                f"₹{row['opening_balance']:,.0f}",
                f"₹{row['total_sales']:,.0f}",
                str(row['order_count']),
                closing,
                diff,
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        
        doc.build(elements)
        return buffer.getvalue()


class ExcelExporter:
    """Export reports to Excel using openpyxl."""
    
    @staticmethod
    def _style_header(ws, row=1):
        """Apply header styling."""
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = colors
        for cell in ws[row]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    @staticmethod
    def export_daily_sales(data: list, start_date: date, end_date: date) -> bytes:
        """Export daily sales to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Sales"
        
        # Headers
        headers = ['Date', 'Total Sales', 'Orders', 'Avg Order', 'Tax', 'Discount']
        ws.append(headers)
        
        # Data
        for row in data:
            ws.append([
                row['date'],
                float(row['total_sales']),
                row['total_orders'],
                float(row['average_order_value']),
                float(row['total_tax']),
                float(row['total_discount']),
            ])
        
        # Totals
        totals_row = len(data) + 2
        ws.cell(row=totals_row, column=1, value='TOTAL')
        ws.cell(row=totals_row, column=2, value=f'=SUM(B2:B{totals_row-1})')
        ws.cell(row=totals_row, column=3, value=f'=SUM(C2:C{totals_row-1})')
        
        # Format
        for col in ['B', 'D', 'E', 'F']:
            for row in range(2, totals_row + 1):
                ws[f'{col}{row}'].number_format = '₹#,##0.00'
        
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    @staticmethod
    def export_product_sales(data: list, start_date: date, end_date: date) -> bytes:
        """Export product sales to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Sales"
        
        headers = ['Product', 'Category', 'Quantity Sold', 'Total Sales', 'Percentage']
        ws.append(headers)
        
        for row in data:
            ws.append([
                row['product_name'],
                row['category_name'],
                float(row['quantity_sold']),
                float(row['total_sales']),
                float(row['percentage']),
            ])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    @staticmethod
    def export_session_summary(data: list, start_date: date, end_date: date) -> bytes:
        """Export session summary to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Session Summary"
        
        headers = [
            'Terminal', 'User', 'Opened At', 'Closed At',
            'Opening Balance', 'Total Sales', 'Order Count',
            'Cash In', 'Cash Out', 'Expected Cash', 'Closing Balance', 'Difference'
        ]
        ws.append(headers)
        
        for row in data:
            ws.append([
                row['terminal_name'],
                row['user_name'],
                row['opened_at'],
                row['closed_at'],
                float(row['opening_balance']),
                float(row['total_sales']),
                row['order_count'],
                float(row['cash_in']),
                float(row['cash_out']),
                float(row['expected_cash']),
                float(row['closing_balance']) if row['closing_balance'] else None,
                float(row['difference']) if row['difference'] else None,
            ])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    @staticmethod
    def export_category_sales(data: list, start_date: date, end_date: date) -> bytes:
        """Export category sales to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Category Sales"
        
        headers = ['Category', 'Total Sales', 'Item Count', 'Percentage']
        ws.append(headers)
        
        for row in data:
            ws.append([
                row['category_name'],
                float(row['total_sales']),
                row['item_count'],
                float(row['percentage']),
            ])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    @staticmethod
    def export_payment_methods(data: list, start_date: date, end_date: date) -> bytes:
        """Export payment method breakdown to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Payment Methods"
        
        headers = ['Payment Method', 'Total Amount', 'Transaction Count', 'Percentage']
        ws.append(headers)
        
        for row in data:
            ws.append([
                row['payment_method_name'],
                float(row['total_amount']),
                row['transaction_count'],
                float(row['percentage']),
            ])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    @staticmethod
    def export_staff_performance(data: list, start_date: date, end_date: date) -> bytes:
        """Export staff performance to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff Performance"
        
        headers = ['Staff Name', 'Total Sales', 'Order Count', 'Avg Order', 'Hours Worked']
        ws.append(headers)
        
        for row in data:
            ws.append([
                row['user_name'],
                float(row['total_sales']),
                row['order_count'],
                float(row['average_order_value']),
                float(row['total_hours_worked']),
            ])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
